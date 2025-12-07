import os, time, pickle, csv, re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import StaleElementReferenceException

COOKIE_FILE = "trip_cookies.pkl"
OUTPUT_FILE = "trip_flights_data.xlsx"

def get_date(days_ahead):
    return (datetime.today() + timedelta(days=days_ahead)).strftime("%Y-%m-%d")

def save_cookies(driver):
    pickle.dump(driver.get_cookies(), open(COOKIE_FILE, "wb"))

def load_cookies(driver):
    cookies = pickle.load(open(COOKIE_FILE, "rb"))
    for cookie in cookies:
        if isinstance(cookie.get("expiry"), float):
            cookie["expiry"] = int(cookie["expiry"])
        driver.add_cookie(cookie)

COLUMNS = [
    "date",
    "airline",
    "departure_time",
    "arrival_time",
    "duration",
    "transit",
    "price",  # overwritten in modal
    "origin",
    "destination",
    "fare_type",
    "baggage",
    "scraped_time",
]

def save_to_excel(data, filename=OUTPUT_FILE):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(COLUMNS)  # write header

    for row in data:
        ws.append([row.get(col, "") for col in COLUMNS])

    wb.save(filename)


def scroll_to_load_flights(driver, pause_time=2, max_scroll=10):
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(max_scroll):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

def get_text_after_scroll(driver, parent, selector):
    try:
        driver.execute_script("arguments[0].scrollIntoView(true);", parent)
        time.sleep(0.5)
        elem = parent.find_element(By.CSS_SELECTOR, selector)
        return elem.get_attribute("textContent").strip()
    except:
        return None

def scrape_fare_modal(driver, f, base_info):
    fares = []
    try:
        # --- Find and click fare button ---
        try:
            button = f.find_element(By.CSS_SELECTOR, "button[data-testid='u_select_btn']")
        except:
            button = f.find_element(By.CSS_SELECTOR, "button.c-result-operate__btn")

        driver.execute_script("arguments[0].scrollIntoView(true);", button)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", button)

        # --- Wait for modal ---
        modal = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".flt-page-modal-wrapper"))
        )
        print("✅ Modal appeared")

        # --- Get first fare card ---
        card = WebDriverWait(modal, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".result-item-flex__wrapper"))
        )

        fare_type, price, baggage = None, None, None

        # --- Fare type ---
        try:
            elem = card.find_element(By.CSS_SELECTOR, "span[data-testid='mid_result_item_class_cabin']")
            fare_type = elem.get_attribute("aria-label").strip()
        except Exception as e:
            print(f"⚠️ No fare type found: {e}")

        # --- Price (from data-price attribute) ---
        try:
            span_price = card.find_element(By.CSS_SELECTOR, "span.o-price-flight")
            price_attr = span_price.get_attribute("data-price")
            if price_attr:
                price = round(float(price_attr))
            else:
                print("⚠️ Missing data-price attribute")
        except Exception as e:
            print(f"⚠️ Price not found: {e}")

        # --- Baggage (Checked baggage from second .carryon.is-ADVANTAGE div) ---
        try:
            carryons = modal.find_elements(By.CSS_SELECTOR, "div.carryon.is-ADVANTAGE")
            print(f"Found {len(carryons)} .carryon.is-ADVANTAGE elements")

            if len(carryons) >= 2:
                target = carryons[1]  # second one
                driver.execute_script("arguments[0].scrollIntoView(true);", target)
                time.sleep(0.5)

                # Get <b> text inside title-content
                baggage_elem = target.find_element(By.CSS_SELECTOR, ".subject .title-content b")
                baggage_text = baggage_elem.text.strip()

                if not baggage_text:
                    # Fallback to span text if b is empty
                    baggage_text = target.find_element(By.CSS_SELECTOR, ".subject .title-content").text.strip()

                match = re.search(r"(\d+)\s*[Kk][Gg]", baggage_text)
                if match:
                    baggage = int(match.group(1))
                    print(f"✅ Parsed baggage: {baggage} KG")
                else:
                    print(f"⚠️ Couldn't parse baggage text: '{baggage_text}'")
                    baggage = 10
            else:
                print("⚠️ Less than 2 carryon.is-ADVANTAGE divs found")
                baggage = 10

        except Exception as e:
            print(f"⚠️ Baggage not found: {e}")
            baggage = 10

        # --- Merge result ---
        fare_data = base_info.copy()
        fare_data.update({
            "fare_type": fare_type,
            "price": price,
            "baggage": baggage,
        })
        fares.append(fare_data)

        print(f"✅ Fare parsed: type={fare_type}, price={price}, baggage={baggage}")

        # --- Close modal ---
        try:
            close_buttons = driver.find_elements(By.CSS_SELECTOR, "i.fi-icon.fi-icon_close_line")
            if close_buttons:
                driver.execute_script("arguments[0].click();", close_buttons[-1])
                time.sleep(2)
                print("✅ Closed the modal")
                # Wait for flight list to refresh
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-flight-id]"))
                )
            else:
                print("❌ No close button found inside modal")
        except Exception as e:
            print(f"⚠️ Could not close modal properly: {e}")

    except Exception as e:
        print(f"❌ Modal scraping failed: {e}")

    return fares



def scrape_flights(origin, dest, date, driver):
    url = f"https://id.trip.com/flights/showfarefirst?dcity={origin}&acity={dest}&ddate={date}&triptype=ow&class=c&quantity=1&locale=en-ID&curr=IDR"
    driver.get(url)

    scroll_to_load_flights(driver)

    try:
        flights = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-flight-id]"))
        )
    except:
        print(f"No flights found for {origin}->{dest} {date}")
        return []

    results = []
    seen_ids = set()

    for index, f in enumerate(flights):
        try:
            # Try to get flight id safely
            try:
                flight_id = f.get_attribute("data-flight-id")
            except StaleElementReferenceException:
                # Re-find all flights and get the current one again
                flights = driver.find_elements(By.CSS_SELECTOR, "div[data-flight-id]")
                if index >= len(flights):
                    print(f"⚠️ Flight {index} disappeared, skipping")
                    continue
                f = flights[index]
                flight_id = f.get_attribute("data-flight-id")

            if not flight_id or flight_id in seen_ids:
                continue
            seen_ids.add(flight_id)

            # Basic info
            airline = get_text_after_scroll(driver, f, ".flights-name")

            # Get origin and destination airport codes (with terminal)
            try:
                codes = f.find_elements(By.CSS_SELECTOR, "span.flight-info-stop__code_e162")
                origin_code = codes[0].text.strip() if len(codes) > 0 else None
                dest_code = codes[1].text.strip() if len(codes) > 1 else None
            except:
                origin_code, dest_code = None, None

            try:
                time_elems = f.find_elements(By.CSS_SELECTOR, ".flight-info-airline__timers_39aa .time_cbcc span")
                dep_time = time_elems[0].text.strip() if len(time_elems) > 0 else None
                arr_time = time_elems[1].text.strip() if len(time_elems) > 1 else None
            except:
                dep_time, arr_time = None, None

            try:
                duration = f.find_element(By.CSS_SELECTOR, ".flight-info-duration_576d span").text.strip()
            except:
                duration = None



            try:
                stop_elem = f.find_element(
                    By.CSS_SELECTOR,
                    "span.flight-info-stop__text_3ee2.dash_text_fb9a"
                )
                stop_text = stop_elem.text.strip()
                print("Raw stop text:", stop_text)

                import re
                stops = None

                stop_text_lower = stop_text.lower()

                if "nonstop" in stop_text_lower or "direct" in stop_text_lower:
                    stops = "Direct"
                else:
                    # Try to find numeric stop info, e.g. "1 stop", "2 stops"
                    match = re.search(r"(\d+)\s+stop", stop_text_lower)
                    if match:
                        num_stops = int(match.group(1))
                        stops = f"{num_stops} stop" if num_stops == 1 else f"{num_stops} stops"
                    elif "in" in stop_text_lower:
                        # Handle "5h 55min in Jakarta" → 1 stop
                        stops = "1 stop"
                    else:
                        stops = None

            except:
                print("fail to get stops")
                stops = None

            base_info = {
                "date": date,
                "airline": airline,
                "departure_time": dep_time,
                "arrival_time": arr_time,
                "duration": duration,
                "transit": stops,
                "price": None,  # overwritten in modal
                "origin": origin_code,
                "destination": dest_code,
                "fare_type": None,
                "baggage": None,
                "scraped_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            # Scrape modal fares
            fares = scrape_fare_modal(driver, f, base_info)

            if fares:
                # Only take the first fare card (avoid duplicates)
                results.append(fares[0])
            else:
                # Fallback to base_info if modal scraping failed
                results.append(base_info)

        except Exception as e:
            print("Error parsing flight:", e)

    return results

ROUTES = [
    # Domestic pairs
    # ("sub", "dps"),
    # ("dps", "sub"),
    # ("jkt", "dps"),
    # ("dps", "jkt"),

    # ("sub", "upg"),
    # ("upg", "sub")

    # ("sub", "sin"),
    # ("sin", "sub")

    # ("sub", "jog"),
    # ("jog", "sub")

    #International
    # ("jkt", "sin"),
    # ("sin", "jkt")

    ("jkt", "kul"),
    # ("kul", "jkt")

    # ("sub", "tyo"),
    # ("tyo", "sub")
]

if __name__ == "__main__":
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://id.trip.com/")

    if os.path.exists(COOKIE_FILE):
        print("Loading cookies...")
        driver.delete_all_cookies()
        load_cookies(driver)
        driver.refresh()
    else:
        print("Please log in manually...")
        time.sleep(60)
        save_cookies(driver)
        print("Cookies saved! Restart the script next time.")

    try:
        for origin, dest in ROUTES:
            for d in range(1, 2):  # tomorrow
                date = get_date(d)
                print(f"Scraping {origin} -> {dest} on {date}")
                data = scrape_flights(origin, dest, date, driver)
                if data:
                    save_to_excel(data)
                print(f"Done: {len(data)} flights")
    finally:
        time.sleep(5)
        driver.quit()
