import time
import re
import os
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

# ========================= CONFIG =========================
ROUTES = [
    # 15 flights = business dan economy saja
    # ("jkt", "sin"),
    # ("sin", "jkt"),
    # ("sub", "dps"),
    # ("dps", "sub"),
    # ("sub", "sin"),
    # ("sin", "sub"),
    # ("sub", "srg"),
    # ("srg", "sub"),
    ("sub", "jkt"),
    ("jkt", "sub"),
]

START_DATE = "2025-10-27"
END_DATE = "2025-11-06"

MAX_FLIGHTS_PER_PAGE = 3
SCROLL_DELAY = 3
OUTPUT_FILE = "tiket_flights_data.xlsx"
# ==========================================================


def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver


def parse_duration(duration_text):
    duration_text = duration_text.lower().replace("j", "h")
    match = re.match(r"(\d+)h(?:\s*(\d+)m)?", duration_text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2)) if match.group(2) else 0
        total_hours = hours + minutes / 60
        return f"{hours}h {minutes}m", round(total_hours, 2)
    return duration_text, None


def extract_baggage_info(fare_text):
    """Parse baggage info into numeric kg value only."""
    fare_text = fare_text.lower()
    match_pc = re.search(r"(\d+)\s*pc", fare_text)
    if match_pc:
        return int(match_pc.group(1)) * 10
    match_kg = re.findall(r"(\d+)\s*kg", fare_text)
    if match_kg:
        val = int(match_kg[-1])
        return val if val > 0 else 10
    return 0


def scrape_flights(origin, dest, date, driver):
    URL = (
        f"https://www.tiket.com/pesawat/search?"
        f"d={origin.upper()}C&a={dest.upper()}C&date={date}&adult=1&child=0&infant=0"
        f"&class=business&dType=CITY&aType=CITY&dLabel={origin}&aLabel={dest}&type=depart&flexiFare=true"
    )
    driver.get(URL)
    wait = WebDriverWait(driver, 30)
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.FlightCard_card_body_wrapper___1uCh")))
    time.sleep(3)

    # Scroll until enough flights are loaded or no more new cards
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        cards = driver.find_elements(By.CSS_SELECTOR, "div.FlightCard_card_body_wrapper___1uCh")
        if len(cards) >= MAX_FLIGHTS_PER_PAGE:
            break
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_DELAY)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    cards = driver.find_elements(By.CSS_SELECTOR, "div.FlightCard_card_body_wrapper___1uCh")
    print(f"✅ Found {len(cards)} flight cards for {origin.upper()} → {dest.upper()} ({date})")

    if len(cards) > MAX_FLIGHTS_PER_PAGE:
        cards = cards[:MAX_FLIGHTS_PER_PAGE]
        print(f"⚙️ Limiting to {MAX_FLIGHTS_PER_PAGE} flight cards per page")

    flights = []
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i in range(1, len(cards) + 1):
        try:
            card = driver.find_element(
                By.XPATH, f"(//div[contains(@class,'FlightCard_card_body_wrapper___1uCh')])[{i}]"
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", card)
            time.sleep(0.5)
            text = card.text.strip()
        except Exception as e:
            print(f"⚠️ Skipped card {i}, reason: {e}")
            continue

        lines = [l.strip() for l in text.split("\n") if l.strip()]
        lines = [l for l in lines if not ("Setelah cashback" in l or "Bisa reschedule" in l)]
        if len(lines) < 8:
            print(f"⚠️ Skipped card {i}, structure unexpected: {lines}")
            continue

        airline = lines[0]
        dep_time = lines[1]
        dep_airport = lines[2]
        duration_text, _ = parse_duration(lines[3])
        transit = 0 if "Langsung" in lines[4] else 1
        arr_time = lines[5]
        arr_airport = lines[6]

        # Find price line containing "IDR"
        price_line = next((l for l in lines if "idr" in l.lower()), None)
        price = re.sub(r"[^\d]", "", price_line) if price_line else ""

        # Default values
        fare_type = "Unknown"
        baggage_kg = 0

        try:
            driver.execute_script("arguments[0].click();", card)
            wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.FareCard_box__EnNnm")))
            time.sleep(2)

            fare_cards = driver.find_elements(By.CSS_SELECTOR, "div.FareCard_box__EnNnm")
            if fare_cards:
                fare_text = fare_cards[0].text.replace("\n", " ").strip()
                baggage_kg = extract_baggage_info(fare_text)
                fare_match = re.search(r"(Ekonomi|Bisnis)", fare_text, re.IGNORECASE)
                fare_type = "Economy" if not fare_match or "eko" in fare_match.group(1).lower() else "Business"

            driver.back()
            wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.FlightCard_card_body_wrapper___1uCh")))
            time.sleep(2)

        except Exception as e:
            print(f"⚠️ Failed to open fare detail for {airline}: {e}")
            try:
                driver.back()
                wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.FlightCard_card_body_wrapper___1uCh")))
                time.sleep(2)
            except:
                pass

        flights.append([
            date, airline, dep_time, arr_time, duration_text, transit,
            price, origin.upper(), dest.upper(), fare_type, baggage_kg, scraped_at
        ])

        print(f"✈️ {airline} | {dep_time} → {arr_time} | {duration_text} | {fare_type} | Bag: {baggage_kg}kg | {price}")

        if len(flights) >= MAX_FLIGHTS_PER_PAGE:
            print(f"🛑 Reached MAX_FLIGHTS_PER_PAGE limit ({MAX_FLIGHTS_PER_PAGE}). Stopping early.")
            break

    return flights


def save_to_excel(data, filename=OUTPUT_FILE):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        print("📎 Appending to existing Excel file...")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Flights"
        ws.append([
            "date", "airline", "departure_time", "arrival_time", "duration",
            "transit", "price_IDR", "origin", "destination", "seat_class",
            "baggage_kg", "scraped_at"
        ])
        print("🆕 Creating new Excel file...")

    for row in data:
        ws.append(row)

    wb.save(filename)
    print(f"✅ Saved (appended) to {filename}")


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)


# ========================= MAIN =========================
driver = init_driver()
start_date = datetime.strptime(START_DATE, "%Y-%m-%d").date()
end_date = datetime.strptime(END_DATE, "%Y-%m-%d").date()

for origin, dest in ROUTES:
    for single_date in daterange(start_date, end_date):
        date_str = single_date.strftime("%Y-%m-%d")
        print(f"\n🚀 Scraping {origin.upper()} → {dest.upper()} on {date_str}")

        flights = scrape_flights(origin, dest, date_str, driver)

        if flights:
            save_to_excel(flights)  # ✅ Save after each page scraped
        else:
            print(f"⚠️ No flights found for {origin.upper()} → {dest.upper()} on {date_str}")

print("\n🎉 All routes finished!")
input("\nPress ENTER to quit...")
driver.quit()
