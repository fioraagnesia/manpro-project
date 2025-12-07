import pickle, os, time, re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

COOKIE_FILE = "trip_cookies.pkl"
OUTPUT_FILE = "trip_hotels_data.xlsx"

city_ids = {
    # per hari 15
    "Surabaya": 1244,
    # "Jakarta": 524,
    # "Bali": 723,
    # "Bandung": 740,
    # "Yogyakarta": 741,
    # "Semarang": 1488,
    # malang
    # "Singapura": 73,
    # "Kuala Lumpur": 315,
    # "Bangkok": 359
}

# ========== DATE RANGE FUNCTION ==========
def get_date_range(days_ahead, stay_length=2):
    checkin = (datetime.today() + timedelta(days=days_ahead)).strftime("%Y-%m-%d")
    checkout = (datetime.today() + timedelta(days=days_ahead + stay_length)).strftime("%Y-%m-%d")
    return checkin, checkout

# ========== COOKIE HANDLERS ==========
def save_cookies(driver):
    pickle.dump(driver.get_cookies(), open(COOKIE_FILE, "wb"))

def load_cookies(driver):
    cookies = pickle.load(open(COOKIE_FILE, "rb"))
    for cookie in cookies:
        if isinstance(cookie.get("expiry"), float):
            cookie["expiry"] = int(cookie["expiry"])
        driver.add_cookie(cookie)

# ========== EXCEL SAVER ==========
def save_to_excel(data, filename=OUTPUT_FILE):
    if not data:
        return

    # If file exists, open it; otherwise create new workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Write header row
        ws.append(list(data[0].keys()))

    # Write each record
    for row in data:
        ws.append(list(row.values()))

    wb.save(filename)
    print(f"✅ Saved {len(data)} records to {filename}")

# ========== SCROLLING ==========
def scroll_limited(driver, max_scrolls=5, pause=2):
    """Scroll a few times to load more hotels."""
    for i in range(max_scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause)

# ========== SCRAPER ==========
def scrape_hotels(city, city_id, checkin, checkout, driver, limit=60):
    url = f"https://id.trip.com/hotels/list?city={city_id}&checkin={checkin}&checkout={checkout}"
    driver.get(url)

    scroll_limited(driver)

    try:
        # Try the modern layout first (div-based)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.hotel-list"))
        )
        hotels = driver.find_elements(By.CSS_SELECTOR, "div.hotel-list > div")
        layout = "div.hotel-list"

    except:
        # Fallback to the <ul> layout
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "ul.long-list.long-list-v8.version-b"))
        )
        hotels = driver.find_elements(By.CSS_SELECTOR, "ul.long-list.long-list-v8.version-b > li")
        layout = "ul.long-list"

    print(f"✅ Detected layout: {layout}")
    print(f"Found {len(hotels)} hotel cards")

    results = []
    for hotel in hotels[:limit]:
        try:

            # Hotel name and URL
            try:
                # Try the modern layout first
                name_elem = hotel.find_element(By.CSS_SELECTOR, ".list-card-title a.name")
                name = name_elem.text.strip()
                url = name_elem.get_attribute("href")
            except:
                try:
                    # Fallback to simpler layout
                    name_elem = hotel.find_element(By.CSS_SELECTOR, ".hotelName")
                    name = name_elem.text.strip()
                    url = None  # this version usually doesn’t have a link
                except:
                    name, url = None, None

            # Star rating
            try:
                stars = len(hotel.find_elements(By.CSS_SELECTOR, ".more-star-repeat i"))
            except:
                stars = None

            # Price
            try:
                price_elem = hotel.find_element(By.CSS_SELECTOR, "p.price-explain")
                price_digits = re.sub(r"[^\d]", "", price_elem.text.strip())
                price = int(price_digits) if price_digits else None
                available = "Yes"
            except:
                price, available = None, "No"

            # Guest rating
            try:
                # --- Guest Rating (using if-else) ---
                if hotel.find_elements(By.CSS_SELECTOR, ".comment-score .real"):
                    guest_rating = hotel.find_element(By.CSS_SELECTOR, ".comment-score .real").text.strip()
                elif hotel.find_elements(By.CSS_SELECTOR, ".score .real"):
                    guest_rating = hotel.find_element(By.CSS_SELECTOR, ".score .real").text.strip()
                else:
                    guest_rating = None
                    # 🧠 Debug: print small snippet of hotel card
                    print("⚠️ No guest rating found in this hotel. HTML snippet:")
                    snippet = hotel.get_attribute("outerHTML")[:500]
                    print(snippet)
            except Exception as e:
                print(f"⚠️ Error finding guest rating: {e}")
                guest_rating = None

            results.append({
                "hotel_name": name,
                "price": price,
                "city": city,
                "country": "Indonesia",
                "hotel_star": stars,
                "guest_rating": guest_rating,
                "checkin_date": checkin,
                "checkout_date": checkout,
                "source_url": url,
            })
        except Exception as e:
            print("⚠️ Error parsing hotel:", e)
            continue

    print(f"✅ Parsed {len(results)} hotels for {city}")
    return results

# ========== MAIN ==========
if __name__ == "__main__":
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://id.trip.com/")

    if os.path.exists(COOKIE_FILE):
        print("Loading saved cookies...")
        driver.delete_all_cookies()
        load_cookies(driver)
        driver.refresh()
    else:
        print("Please log in manually...")
        time.sleep(60)
        save_cookies(driver)
        print("Cookies saved! Restart the script next time.")

    try:
        for days_ahead in range(8, 18):  # adjust range if needed
            checkin, checkout = get_date_range(days_ahead, 1)
            for city, city_id in city_ids.items():
                print(f"\n🔎 Scraping {city} {checkin} → {checkout}")
                data = scrape_hotels(city, city_id, checkin, checkout, driver, limit=15)
                if data:
                    save_to_excel(data)
                print(f"✅ Done: {len(data)} hotels from {city}")
    finally:
        driver.quit()
